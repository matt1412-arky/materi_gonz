from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import os

folder_path = r"C:\Users\Lenovo\OneDrive\Documents\materi_gonz\kelas 11\Otomasi\Laporan"
os.makedirs(folder_path, exist_ok=True)

file_path = os.path.join(folder_path, "Laporan_Nilai.xlsx")
def create_report(data, filename):
    """
    Buat laporan Excel otomatis dengan format dan chart
    
    data = {
        'Nama': ['Ali', 'Budi', 'Citra'],
        'Matematika': [85, 78, 92],
        'Fisika': [80, 85, 88],
        'Kimia': [88, 76, 90]
    }
    """
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Nilai"
    
    # Header
    ws['A1'] = "LAPORAN NILAI SISWA"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A1:E1')
    
    ws['A2'] = f"Tanggal: {datetime.now().strftime('%d-%m-%Y')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A2:E2')
    
    # Header tabel (baris 4)
    headers = ['No', 'Nama', 'Matematika', 'Fisika', 'Kimia', 'Rata-rata']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data siswa
    nama_list = data['Nama']
    for i, nama in enumerate(nama_list, start=1):
        row = i + 4  # Mulai dari baris 5
        
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=nama)
        ws.cell(row=row, column=3, value=data['Matematika'][i-1])
        ws.cell(row=row, column=4, value=data['Fisika'][i-1])
        ws.cell(row=row, column=5, value=data['Kimia'][i-1])
        
        # Rata-rata dengan formula
        ws.cell(row=row, column=6, value=f"=AVERAGE(C{row}:E{row})")
        ws.cell(row=row, column=6).number_format = '0.00'
    
    # Tambah baris total rata-rata
    last_row = len(nama_list) + 5
    ws.cell(row=last_row, column=1, value="RATA-RATA KELAS")
    ws.cell(row=last_row, column=1).font = Font(bold=True)
    ws.merge_cells(f'A{last_row}:B{last_row}')
    
    for col in range(3, 7):
        ws.cell(row=last_row, column=col, value=f"=AVERAGE({chr(64+col)}5:{chr(64+col)}{last_row-1})")
        ws.cell(row=last_row, column=col).number_format = '0.00'
        ws.cell(row=last_row, column=col).font = Font(bold=True)
        ws.cell(row=last_row, column=col).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    # Atur lebar kolom
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # Buat chart
    chart = BarChart()
    chart.title = "Grafik Nilai Siswa"
    chart.x_axis.title = "Siswa"
    chart.y_axis.title = "Nilai"
    
    # Data untuk chart
    data_ref = Reference(ws, min_col=3, min_row=4, max_col=5, max_row=last_row-1)
    categories = Reference(ws, min_col=2, min_row=5, max_row=last_row-1)
    
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories)
    
    # Posisi chart
    ws.add_chart(chart, f"H4")
    
    # Save
    wb.save(filename)
    print(f"✅ Laporan berhasil dibuat: {filename}")

# Cara pakai
if __name__ == "__main__":
    data_siswa = {
        'Nama': ['Ali Rahman', 'Budi Santoso', 'Citra Dewi', 'Dina Maharani', 'Eko Prasetyo'],
        'Matematika': [85, 78, 92, 88, 75],
        'Fisika': [80, 85, 89, 82, 78],
        'Kimia': [88, 76, 94, 85, 80]
    }
    
    create_report(data_siswa, file_path)