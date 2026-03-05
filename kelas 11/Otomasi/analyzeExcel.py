from openpyxl import load_workbook
import statistics
import os

def analyze_excel(filename):
    """
    Analisis data dari Excel
    """
    
    # Membuka file Excel
    wb = load_workbook(filename)
    ws = wb.active
    
    print("📊 ANALISIS DATA NILAI")
    print("="*50)
    
    # Baca data (skip header)
    data = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row-1, min_col=3, max_col=5, values_only=True):
        if all(isinstance(val, (int, float)) for val in row):
            data.append(row)
    
    if not data:
        print("❌ Tidak ada data untuk dianalisis")
        return
    
    # Analisis per mata pelajaran
    mapel = ['Matematika', 'Fisika', 'Kimia']
    
    for i, nama_mapel in enumerate(mapel):
        nilai_mapel = [row[i] for row in data]
        
        print(f"\n📖 {nama_mapel}:")
        print(f"   Rata-rata: {statistics.mean(nilai_mapel):.2f}")
        print(f"   Tertinggi: {max(nilai_mapel)}")
        print(f"   Terendah: {min(nilai_mapel)}")
        print(f"   Median: {statistics.median(nilai_mapel):.2f}")
        
        # Hitung siswa lulus (≥75)
        lulus = sum(1 for n in nilai_mapel if n >= 75)
        persen_lulus = (lulus / len(nilai_mapel)) * 100
        print(f"   Lulus: {lulus}/{len(nilai_mapel)} ({persen_lulus:.1f}%)")


# ===============================
# PATH FILE EXCEL
# ===============================
folder_path = r"C:\Users\Lenovo\OneDrive\Documents\materi_gonz\kelas 11\Otomasi\Laporan"

# memastikan folder ada
os.makedirs(folder_path, exist_ok=True)

# membuat path file
file_path = os.path.join(folder_path, "Laporan_Nilai.xlsx")

# menjalankan analisis
analyze_excel(file_path)