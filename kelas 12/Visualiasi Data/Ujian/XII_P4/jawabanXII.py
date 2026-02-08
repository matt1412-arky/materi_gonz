import openpyxl
import matplotlib.pyplot as plt
import os

# =========================
# BUAT FOLDER OUTPUT
# =========================

output_folder = "C:/Users/Lenovo/OneDrive/Documents/materi_gonz/kelas 12/Visualiasi Data/Ujian/XII_P4/hasil_grafik"

if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# =========================
# LOAD EXCEL
# =========================

wb = openpyxl.load_workbook(
    "C:/Users/Lenovo/OneDrive/Documents/materi_gonz/kelas 12/Visualiasi Data/Ujian/dataset/dataset_visualisasi_sekolah.xlsx"
)

# =========================
# 1. Aktivitas_Siswa (Pie)
# =========================

sheet = wb["Aktivitas_Siswa"]

aktivitas = []
jumlah = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    aktivitas.append(row[0])
    jumlah.append(row[1])

plt.figure()
plt.pie(jumlah, labels=aktivitas, autopct="%1.1f%%")
plt.title("Distribusi Aktivitas Siswa")
plt.tight_layout()
plt.savefig(output_folder + "/aktivitas.png")
plt.close()

# =========================
# 2. Nilai_Ujian (Histogram)
# =========================

sheet = wb["Nilai_Ujian"]

nilai = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    nilai.append(row[0])

plt.figure()
plt.hist(nilai, bins=8)
plt.title("Distribusi Nilai Ujian")
plt.xlabel("Nilai")
plt.ylabel("Jumlah Siswa")
plt.tight_layout()
plt.savefig(output_folder + "/nilai.png")
plt.close()

# =========================
# 3. Tinggi_Berat (Scatter)
# =========================

sheet = wb["Tinggi_Berat"]

tinggi = []
berat = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    tinggi.append(row[0])
    berat.append(row[1])

plt.figure()
plt.scatter(tinggi, berat)
plt.title("Hubungan Tinggi dan Berat Badan")
plt.xlabel("Tinggi (cm)")
plt.ylabel("Berat (kg)")
plt.tight_layout()
plt.savefig(output_folder + "/tinggi_berat.png")
plt.close()

# =========================
# 4. Jajan_Harian (Histogram)
# =========================

sheet = wb["Jajan_Harian"]

jajan = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    jajan.append(row[0])

plt.figure()
plt.hist(jajan, bins=7)
plt.title("Distribusi Uang Jajan Harian")
plt.xlabel("Rupiah")
plt.ylabel("Jumlah Siswa")
plt.tight_layout()
plt.savefig(output_folder + "/jajan.png")
plt.close()

# =========================
# 5. Transport (Pie)
# =========================

sheet = wb["Transport"]

transport = []
jumlah = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    transport.append(row[0])
    jumlah.append(row[1])

plt.figure()
plt.pie(jumlah, labels=transport, autopct="%1.1f%%")
plt.title("Transportasi Siswa ke Sekolah")
plt.tight_layout()
plt.savefig(output_folder + "/transport.png")
plt.close()

print("SEMUA 5 DATASET BERHASIL DIVISUALISASIKAN.")